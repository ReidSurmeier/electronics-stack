"""BOM sourcing health checker — v2.

Routing pipeline eliminates ~85% false-positives from anti-scrape blocks:
  1. Search-placeholder URLs (Amazon /s?k=, AliExpress wholesale) → marked OK, not fetched.
  2. Digikey product URLs → Digikey API v4 (if creds present).
  3. Mouser product URLs  → Mouser API v2 (if key present).
  4. LCSC product URLs    → local jlcparts SQLite (if DB cached).
  5. Everything else      → HTTP HEAD with realistic UA rotation.
  6. 403/502/503          → Browserbase escalation (capped at 30/BOM, cached 7d).

Usage:
    sourcing_health.py <BOM.xlsx> [--with-api]

Return dict keys (backward-compat + v2 additions):
    xlsx, rows_checked, findings, lifecycle,
    by_via, escalations, actionable_failures
"""
from __future__ import annotations

import re
import sys
import time
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from sourcing_routes import (
    BB_ESCALATION_CAP,
    browserbase_fetch,
    cache_get,
    cache_set,
    classify_url,
    http_check,
)


# ---------------------------------------------------------------------------
# URL checker
# ---------------------------------------------------------------------------

_NO_DELAY_VIAS = frozenset({"cached", "search_placeholder",
                             "digikey_api", "mouser_api", "lcsc_db"})


def check_url(
    url: str,
    timeout: float = 10.0,
    escalation_budget: list[int] | None = None,
) -> dict:
    """Check one URL through the full pipeline.

    Args:
        url: URL to check.
        timeout: HTTP timeout seconds.
        escalation_budget: Mutable [remaining] counter; decremented on bb calls.

    Returns:
        Result dict with url, status, code, via, escalated, actionable.
    """
    if escalation_budget is None:
        escalation_budget = [BB_ESCALATION_CAP]

    if not url or not url.startswith(("http://", "https://")):
        return {"url": url, "status": "skip", "code": None,
                "via": "head", "escalated": False, "actionable": False,
                "note": "non-http URL"}

    cached = cache_get(url)
    if cached:
        cached["url"] = url
        return cached

    routed = classify_url(url)
    if routed is not None:
        routed["url"] = url
        cache_set(url, routed)
        return routed

    result = http_check(url, timeout=timeout)
    result["url"] = url

    if result.pop("_needs_escalation", False):
        if escalation_budget[0] > 0:
            escalation_budget[0] -= 1
            bb = browserbase_fetch(url)
            bb["url"] = url
            if bb["status"] == "ok":
                result = bb
            else:
                result.update({"via": "browserbase", "escalated": True,
                               "actionable": True,
                               "note": bb.get("note", "")})
        else:
            result.update({"actionable": False,
                           "status": "unchecked_anti_scrape",
                           "note": "escalation budget exhausted"})

    cache_set(url, result)
    return result


# ---------------------------------------------------------------------------
# BOM walker
# ---------------------------------------------------------------------------

def walk_bom(xlsx_path: str | Path) -> list[dict]:
    """Return row dicts from the Sourcing sheet."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Sourcing" not in wb.sheetnames:
        return []
    ws = wb["Sourcing"]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows


# ---------------------------------------------------------------------------
# Audit
# ---------------------------------------------------------------------------

def audit(xlsx_path: str | Path, with_api: bool = False) -> dict:
    """Audit all sourcing URLs in the BOM.

    Args:
        xlsx_path: Path to xlsx file.
        with_api: If True, also run Digikey keyword lifecycle checks (legacy).

    Returns:
        Dict: xlsx, rows_checked, findings, lifecycle,
              by_via, escalations, actionable_failures.
    """
    rows = walk_bom(xlsx_path)
    findings: list[dict] = []
    lifecycle_findings: list[dict] = []
    url_columns = ["Existing URL", "Amazon URL", "AliExpress URL"]
    escalation_budget = [BB_ESCALATION_CAP]

    for row in rows:
        item = row.get("Item", "")
        for col in url_columns:
            url = str(row.get(col, "") or "").strip()
            if not url or url.startswith("(") or not url.startswith("http"):
                continue
            result = check_url(url, escalation_budget=escalation_budget)
            result["item"] = item
            result["column"] = col
            findings.append(result)
            if result.get("via") not in _NO_DELAY_VIAS:
                time.sleep(0.3)

    if with_api:
        try:
            from digikey_client import DigikeyClient
            dk = DigikeyClient.from_env()
            for row in rows:
                item = str(row.get("Item", ""))
                mpns = re.findall(r"\b[A-Z][A-Z0-9\-/]{3,30}\b", item)
                seen: set[str] = set()
                for mpn in mpns[:2]:
                    if mpn in seen:
                        continue
                    seen.add(mpn)
                    try:
                        out = dk.keyword_search(mpn, limit=1)
                        prods = out.get("Products", [])
                        if not prods:
                            continue
                        p = prods[0]
                        ps = p.get("ProductStatus", {})
                        status = ps.get("Status", "?") if isinstance(ps, dict) else str(ps)
                        stock = p.get("QuantityAvailable", 0)
                        bad = {"Obsolete", "Not For New Designs",
                               "Discontinued at Digi-Key", "Last Time Buy"}
                        if status in bad:
                            lifecycle_findings.append({
                                "item": item, "mpn_searched": mpn,
                                "status": status, "stock": stock,
                                "severity": "HIGH" if status == "Obsolete" else "MEDIUM",
                            })
                    except Exception:
                        pass
                    time.sleep(0.1)
        except (RuntimeError, ImportError):
            pass

    by_via: dict[str, int] = {}
    for f in findings:
        v = f.get("via", "head")
        by_via[v] = by_via.get(v, 0) + 1

    escalations = sum(1 for f in findings if f.get("escalated", False))
    actionable_failures = sum(
        1 for f in findings
        if f.get("actionable") and f.get("status") not in ("ok", "skip", "ok_search_url")
    )

    return {
        "xlsx": str(xlsx_path),
        "rows_checked": len(rows),
        "findings": findings,
        "lifecycle": lifecycle_findings,
        "by_via": by_via,
        "escalations": escalations,
        "actionable_failures": actionable_failures,
    }


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def report(audit_out: dict) -> str:
    lines = [f"Sourcing health: {audit_out['xlsx']}"]
    f = audit_out["findings"]
    ok = sum(1 for x in f if x["status"] in ("ok", "ok_search_url", "skip"))
    actionable = audit_out.get("actionable_failures", 0)
    lines.append(f"  URLs: {len(f)}  OK/placeholder: {ok}  Actionable failures: {actionable}")
    bv = audit_out.get("by_via", {})
    if bv:
        lines.append("  Via: " + "  ".join(f"{k}={v}" for k, v in sorted(bv.items())))
    escs = audit_out.get("escalations", 0)
    if escs:
        lines.append(f"  Browserbase escalations: {escs}")
    for x in f:
        if x.get("actionable") and x["status"] not in ("ok", "skip", "ok_search_url"):
            lines.append(
                f"  [{x['status']:>22s}][{x.get('via','?'):>18s}]"
                f" {str(x.get('item','?'))[:35]:<35s}"
                f" {x['column']:>14s}: {x['url'][:70]}"
            )
    lc = audit_out.get("lifecycle", [])
    if lc:
        lines.append(f"  Lifecycle alerts (Digikey API): {len(lc)}")
        for L in lc:
            lines.append(
                f"    [{L['severity']}] {L['mpn_searched']:<20s}"
                f" {L['status']:<22s} stock={L['stock']:<8}"
                f"  ({L['item'][:40]})"
            )
    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: sourcing_health.py <BOM.xlsx> [--with-api]")
        sys.exit(1)
    with_api = "--with-api" in sys.argv
    out = audit(sys.argv[1], with_api=with_api)
    print(report(out))
    sys.exit(1 if out["actionable_failures"] > 0 else 0)
